#!/usr/bin/env python3
"""Genera un Excel de importación de plan de cuentas ERPNext a partir del
PUCT boliviano (plan_bo_puct.xlsx), filtrado por sector económico."""

import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_FILE = SCRIPT_DIR / "plan_bo_puct.xlsx"
SHEET_NAME = "PUCT"

LEVEL_COLUMNS = ["C", "G", "SG", "CP", "CA"]  # columnas A-E, en orden jerárquico
NAME_COLUMN_INDEX = 6  # columna F

ROOT_TYPE_BY_CLASS = {
    "1": "Asset",
    "2": "Liability",
    "3": "Equity",
    "4": "Income",
    "5": "Expense",
}

# Siglas y conceptos que se mantienen en mayúsculas al formatear nombres.
# Editar/ampliar aquí según se detecten nuevos casos.
ACRONYMS = {
    "IVA", "IT", "IUE", "ICE", "IEHD", "ITF", "NIT", "SIN", "TGN", "BCB",
    "AFP", "UFV", "ONG", "RUC", "RC-IVA", "S.A.", "S.R.L.", "CEDEIM", "CENOCREF",
}

# Reglas heurísticas para detectar Account Type a partir del nombre de la
# cuenta (ya sin tildes y en mayúsculas). Se evalúan en orden; la primera
# que coincide gana. Ampliar según se necesiten más tipos.
ACCOUNT_TYPE_RULES = [
    (re.compile(r"\bCAJA\b"), "Cash"),
    (re.compile(r"\bBANCOS?\b"), "Bank"),
    (re.compile(r"DEPRECIACION ACUMULADA"), "Accumulated Depreciation"),
    (re.compile(r"DEPRECIACION"), "Depreciation"),
    (re.compile(r"\bIVA\b|CREDITO FISCAL|DEBITO FISCAL|RC-IVA|\bIT\b|\bIUE\b"), "Tax"),
    (re.compile(r"EXISTENCIA|INVENTARIO|MERCADERIA"), "Stock"),
    (re.compile(r"POR COBRAR"), "Receivable"),
    (re.compile(r"POR PAGAR"), "Payable"),
    (re.compile(r"ACTIVO FIJO|BIENES DE USO"), "Fixed Asset"),
]

PLACEHOLDER_NAME = "XXX"
DEFAULT_CURRENCY = "BOB"

# Bancos que se agregan como subcuentas reales de "Bancos" (1.1.1.002), que
# en el origen solo trae hijos placeholder. No incluye Banco Fassil por
# estar intervenido/cerrado por ASFI desde 2023. Editar/ampliar aquí.
BANCOS_ACCOUNT_NUMBER = "1.1.1.002"
BOLIVIAN_BANKS = [
    "Banco Nacional de Bolivia (BNB)",
    "Banco Mercantil Santa Cruz",
    "Banco de Crédito de Bolivia (BCP)",
    "Banco Bisa",
    "Banco Unión",
    "Banco Económico",
    "Banco Ganadero",
    "Banco FIE",
    "Banco Solidario (BancoSol)",
    "Banco Fortaleza",
    "Banco Prodem",
]

# Cuentas requeridas por la configuración base de ERPNext (Company / Stock
# Settings / HR Settings), que deben existir en cualquier sector sin importar
# las marcas "X" del PUCT. "new_parent" solo se define para las que hay que
# crear (no existen en el PUCT); las reusadas ya están en el árbol.
ERP_CONFIG_ACCOUNTS = [
    {  # Company: Write Off Account (ERPNext no tiene un Account Type
        # dedicado para esto; queda como "Expense Account" vía heurístico)
        "number": "5.5.1.010",
        "new_parent": "5.5.1",
        "raw_name": "CASTIGOS Y AJUSTES CONTABLES",
        "account_type_override": None,
    },
    {  # Company: Round Off Account (cuenta ya existente, se reusa)
        "number": "5.6.1.004",
        "account_type_override": "Round Off",
    },
    {  # Company: Asset Received But Not Billed
        "number": "2.1.2.024",
        "new_parent": "2.1.2",
        "raw_name": "ACTIVOS FIJOS RECIBIDOS NO FACTURADOS",
        "account_type_override": "Asset Received But Not Billed",
    },
    {  # Company: Default Employee Advance Account
        "number": "1.1.2.020",
        "new_parent": "1.1.2",
        "raw_name": "ANTICIPOS A EMPLEADOS",
        "account_type_override": None,
    },
    {  # HR Settings: Default Payroll Payable Account (cuenta ya existente)
        "number": "2.1.3.001",
        "account_type_override": "Payable",
    },
    {  # Stock Settings: Stock Adjustment Account (cuenta ya existente;
        # override necesario porque el heurístico la detectaría como "Stock")
        "number": "5.1.4.001",
        "account_type_override": "Stock Adjustment",
    },
    {  # Stock Settings: Stock Received But Not Billed
        "number": "2.1.2.025",
        "new_parent": "2.1.2",
        "raw_name": "EXISTENCIAS RECIBIDAS NO FACTURADAS",
        "account_type_override": "Stock Received But Not Billed",
    },
    {  # Stock Settings: Expenses Included In Valuation
        "number": "5.1.1.009",
        "new_parent": "5.1.1",
        "raw_name": "GASTOS INCLUIDOS EN LA VALORACIÓN DE EXISTENCIAS",
        "account_type_override": "Expenses Included In Valuation",
    },
]

ACCOUNT_TYPE_OVERRIDES = {
    entry["number"]: entry["account_type_override"]
    for entry in ERP_CONFIG_ACCOUNTS
    if entry.get("account_type_override")
}

OUTPUT_HEADERS = [
    "Account Name",
    "Parent Account",
    "Account Number",
    "Parent Account Number",
    "Is Group",
    "Account Type",
    "Root Type",
    "Account Currency",
]


def strip_accents(text):
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def norm_code(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def normalize_for_match(text):
    return strip_accents(text).upper()


ACRONYMS_NORMALIZED = {normalize_for_match(a) for a in ACRONYMS}


def format_account_name(raw_name, level, preformatted=False):
    if preformatted:  # nombres propios ya con la capitalización final (ej. bancos)
        return raw_name.strip()
    if level == "C":  # nivel Clase (raíz): siempre en mayúsculas completas
        return raw_name.strip().upper()
    return sentence_case(raw_name)


def sentence_case(raw_name):
    words = [w for w in raw_name.strip().split(" ") if w]
    out_words = []
    for w in words:
        token = normalize_for_match(w).strip(",.;:()")
        if token in ACRONYMS_NORMALIZED:
            out_words.append(w.upper())
        else:
            out_words.append(w.lower())
    result = " ".join(out_words)
    if result:
        result = result[0].upper() + result[1:]
    return result


def detect_account_type(raw_name, root_type, is_group):
    text = normalize_for_match(raw_name)
    for pattern, atype in ACCOUNT_TYPE_RULES:
        if pattern.search(text):
            return atype
    if not is_group:
        if root_type == "Income":
            return "Income Account"
        if root_type == "Expense":
            return "Expense Account"
    return ""


def load_sector_columns(ws):
    """Detecta dinámicamente las columnas de sector a partir de la fila 1,
    a partir de la columna G (índice 7) en adelante."""
    sector_columns = {}
    col = 7
    while True:
        header = ws.cell(row=1, column=col).value
        if header is None or str(header).strip() == "":
            break
        key = normalize_for_match(str(header))
        sector_columns[key] = col
        col += 1
    return sector_columns


def resolve_sector(arg, sector_columns):
    key = normalize_for_match(arg)
    if key in sector_columns:
        return key
    raise SystemExit(
        f"Sector '{arg}' no reconocido. Sectores válidos: "
        + ", ".join(sorted(sector_columns.keys()))
    )


def build_tree(ws, sector_columns):
    """Recorre todas las filas de datos y arma nodes/children_map,
    excluyendo siempre las filas CA placeholder ('XXX')."""
    nodes = {}
    children_map = {}

    max_row = ws.max_row
    for row in range(2, max_row + 1):
        level_codes = [
            norm_code(ws.cell(row=row, column=i + 1).value)
            for i in range(len(LEVEL_COLUMNS))
        ]
        path_parts = [c for c in level_codes if c]
        if not path_parts:
            continue

        raw_name = ws.cell(row=row, column=NAME_COLUMN_INDEX).value
        raw_name = "" if raw_name is None else str(raw_name).strip()

        # nivel = índice del último código no vacío
        level_index = max(i for i, c in enumerate(level_codes) if c)
        level = LEVEL_COLUMNS[level_index]

        if level == "CA" and raw_name.upper() == PLACEHOLDER_NAME:
            continue  # placeholder, se descarta

        number = ".".join(path_parts)
        parent_number = ".".join(path_parts[:-1]) if len(path_parts) > 1 else ""

        class_code = path_parts[0]
        root_type = ROOT_TYPE_BY_CLASS.get(class_code)

        sector_flags = {}
        for sector_key, col in sector_columns.items():
            val = ws.cell(row=row, column=col).value
            sector_flags[sector_key] = bool(val) and str(val).strip().upper() == "X"

        nodes[number] = {
            "level": level,
            "number": number,
            "parent_number": parent_number,
            "raw_name": raw_name,
            "root_type": root_type,
            "sector_flags": sector_flags,
        }
        children_map.setdefault(parent_number, []).append(number)

    return nodes, children_map


def merge_single_ca_children(nodes, children_map):
    """Cuando una Cuenta Principal tiene exactamente un hijo CA real (no
    placeholder), el origen los usa como el mismo concepto repetido (o, si
    el nombre de la CP está en blanco, el nombre real vive solo en la CA).
    En ambos casos se fusionan en una sola cuenta hoja al nivel CP; solo se
    conserva la CA como hijo aparte cuando hay 2+ CA reales distintos
    (desglose genuino, ej. Valores fiscales negociables -> CEDEIM/CENOCREF)."""
    for number, node in list(nodes.items()):
        if node["level"] != "CP":
            continue
        children = children_map.get(number, [])
        if len(children) == 1:
            child = nodes[children[0]]
            if not node["raw_name"]:
                node["raw_name"] = child["raw_name"]
            del nodes[children[0]]
            del children_map[number]


def inject_bank_accounts(nodes, children_map):
    """Agrega los bancos de BOLIVIAN_BANKS como hijos reales de la cuenta
    'Bancos', que en el origen solo trae hijos placeholder ('XXX')."""
    parent = nodes.get(BANCOS_ACCOUNT_NUMBER)
    if parent is None:
        return

    for i, bank_name in enumerate(BOLIVIAN_BANKS, start=1):
        number = f"{BANCOS_ACCOUNT_NUMBER}.{i:03d}"
        nodes[number] = {
            "level": "CA",
            "number": number,
            "parent_number": BANCOS_ACCOUNT_NUMBER,
            "raw_name": bank_name,
            "root_type": parent["root_type"],
            "sector_flags": {},
            "preformatted": True,
        }
        children_map.setdefault(BANCOS_ACCOUNT_NUMBER, []).append(number)


def inject_erp_config_accounts(nodes, children_map):
    """Crea las cuentas de ERP_CONFIG_ACCOUNTS que no existen todavía en el
    PUCT (las que traen 'new_parent'); las que ya existen se dejan tal cual."""
    for entry in ERP_CONFIG_ACCOUNTS:
        number = entry["number"]
        if number in nodes or "new_parent" not in entry:
            continue
        parent_number = entry["new_parent"]
        parent = nodes.get(parent_number)
        if parent is None:
            continue
        nodes[number] = {
            "level": "CP",
            "number": number,
            "parent_number": parent_number,
            "raw_name": entry["raw_name"],
            "root_type": parent["root_type"],
            "sector_flags": {},
        }
        children_map.setdefault(parent_number, []).append(number)


def force_include(nodes, included, numbers):
    """Marca como incluidos estos números y toda su cadena de padres, sin
    importar el sector solicitado (cuentas requeridas por ERPNext base)."""
    for number in numbers:
        n = number
        while n and n in nodes:
            if included.get(n):
                break
            included[n] = True
            n = nodes[n]["parent_number"]


def compute_inclusion(nodes, children_map, sector_key):
    included = {}

    def is_included(number):
        if number in included:
            return included[number]
        node = nodes[number]
        level = node["level"]
        if level == "CP":
            result = node["sector_flags"].get(sector_key, False)
        elif level == "CA":
            parent_number = node["parent_number"]
            result = is_included(parent_number) if parent_number in nodes else False
        else:  # C, G, SG
            own = node["sector_flags"].get(sector_key, False)
            childs = children_map.get(number, [])
            result = own or any(is_included(ch) for ch in childs)
        included[number] = result
        return result

    for number in nodes:
        is_included(number)

    return included


def generate(sector_arg, output_path=None):
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    sector_columns = load_sector_columns(ws)
    sector_key = resolve_sector(sector_arg, sector_columns)

    nodes, children_map = build_tree(ws, sector_columns)
    merge_single_ca_children(nodes, children_map)
    inject_bank_accounts(nodes, children_map)
    inject_erp_config_accounts(nodes, children_map)
    included = compute_inclusion(nodes, children_map, sector_key)
    force_include(nodes, included, [entry["number"] for entry in ERP_CONFIG_ACCOUNTS])

    included_numbers = {n for n, v in included.items() if v}

    cased_names = {
        n: format_account_name(
            nodes[n]["raw_name"], nodes[n]["level"], nodes[n].get("preformatted", False)
        )
        for n in included_numbers
    }

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "plan"
    out_ws.append(OUTPUT_HEADERS)

    for number, node in nodes.items():
        if number not in included_numbers:
            continue

        children = children_map.get(number, [])
        is_group = 1 if any(ch in included_numbers for ch in children) else 0

        parent_number = node["parent_number"]
        parent_name = cased_names.get(parent_number, "") if parent_number in included_numbers else ""

        account_name = cased_names[number]
        account_type = ACCOUNT_TYPE_OVERRIDES.get(number) or detect_account_type(
            node["raw_name"], node["root_type"], is_group
        )

        out_ws.append([
            account_name,
            parent_name,
            number,
            parent_number if parent_number in included_numbers else "",
            is_group,
            account_type,
            node["root_type"],
            DEFAULT_CURRENCY,
        ])

    if output_path is None:
        slug = strip_accents(sector_arg).strip().lower().replace(" ", "_")
        output_path = SCRIPT_DIR / f"plan_bo_{slug}.xlsx"
    else:
        output_path = Path(output_path)

    out_wb.save(output_path)
    return output_path, len(included_numbers)


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 generar_puct_erpnext.py <Sector> [ruta_salida.xlsx]")
        sys.exit(1)

    sector_arg = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    path, count = generate(sector_arg, output_path)
    print(f"Generado: {path} ({count} cuentas)")


if __name__ == "__main__":
    main()
